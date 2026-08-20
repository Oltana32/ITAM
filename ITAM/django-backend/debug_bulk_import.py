import os
import django
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from django.contrib.auth import get_user_model
from django.test.client import RequestFactory
from apps.assets.views import AssetViewSet

path = os.path.abspath('c:/Users/A/Downloads/ITAM/asset_import_template_test.xlsx')
print('Template path:', path)
wb = openpyxl.load_workbook(path)
print('SHEETS:', wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    headers = [c.value if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f'\nSheet: {s}')
    print('Headers:', headers)

User = get_user_model()
user = User.objects.filter(is_superuser=True).first()
if not user:
    user = User.objects.create_superuser('tmpladmin', 'tmpl@example.com', 'templpass')

from django.test import Client
from apps.manufacturers.models import Manufacturer
from apps.locations.models import Location

Manufacturer.objects.update_or_create(name='Example Manufacturer')
Location.objects.update_or_create(name='Main Office')

client = Client()
client.force_login(user)
with open(path, 'rb') as f:
    response = client.post('/api/assets/bulk-import/', {'file': f})

print('\nBulk import response status:', response.status_code)
try:
    print('Response JSON:', response.json())
except Exception:
    print('Response content:', response.content[:2000])
