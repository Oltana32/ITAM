import openpyxl
wb=openpyxl.load_workbook('c:/Users/A/Downloads/ITAM/asset_import_template_test.xlsx')
print('SHEETS:', wb.sheetnames)
for s in wb.sheetnames:
    ws=wb[s]
    try:
        headers=[(c.value if c.value is not None else '') for c in next(ws.iter_rows(min_row=1,max_row=1))]
    except StopIteration:
        headers=[]
    print('\nSheet:', s)
    print('Headers:', headers)
