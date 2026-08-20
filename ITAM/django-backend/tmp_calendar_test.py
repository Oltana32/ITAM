"""Compare date cell XML to find what triggers Excel 365 calendar icon."""
import io
import re
import zipfile
from datetime import date

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2, FORMAT_DATE_XLSX14, builtin_format_id
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.datavalidation import DataValidation

DATE_MIN = str(int(to_excel(date(2000, 1, 1))))
DATE_MAX = str(int(to_excel(date(2100, 12, 31))))


def inspect(label: str, wb: openpyxl.Workbook) -> None:
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    z = zipfile.ZipFile(bio)
    sheet = z.read("xl/worksheets/sheet1.xml").decode()
    styles = z.read("xl/styles.xml").decode()

    cell = re.search(r'<c r="A2"[^>]*/>', sheet)
    print(f"\n=== {label} ===")
    print("A2:", cell.group(0) if cell else "missing")

    dv = re.search(r'<dataValidation[^>]*type="date"[^>]*/>', sheet)
    print("date DV:", dv.group(0) if dv else "missing")

    # show numFmts and cellXfs for style index
    style_idx = None
    if cell:
        m = re.search(r's="(\d+)"', cell.group(0))
        if m:
            style_idx = int(m.group(0).split('"')[1])
    if style_idx is not None:
        xfs = re.findall(r"<xf[^>]+/>", styles)
        if style_idx < len(xfs):
            print("cellXf:", xfs[style_idx])
            m = re.search(r'numFmtId="(\d+)"', xfs[style_idx])
            if m:
                nf_id = m.group(1)
                print("numFmtId:", nf_id, "(builtin 14 = short date)")


def make_wb(apply_style_fn, show_drop_down=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    kwargs = dict(
        type="date",
        operator="between",
        formula1=DATE_MIN,
        formula2=DATE_MAX,
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
    )
    if show_drop_down is not None:
        kwargs["showDropDown"] = show_drop_down
    dv = DataValidation(**kwargs)
    dv.add("A2:A10")
    ws.add_data_validation(dv)
    apply_style_fn(wb, ws)
    return wb


# A: current approach - NamedStyle with yyyy-mm-dd
def style_named(wb, ws):
    name = "itam_import_date"
    if name not in wb.named_styles:
        s = NamedStyle(name=name)
        s.number_format = FORMAT_DATE_YYYYMMDD2
        wb.add_named_style(s)
    for row in range(2, 11):
        ws.cell(row=row, column=1).style = name


# B: direct custom format yyyy-mm-dd
def style_custom(wb, ws):
    for row in range(2, 11):
        ws.cell(row=row, column=1).number_format = FORMAT_DATE_YYYYMMDD2


# C: built-in short date format (mm-dd-yy)
def style_builtin(wb, ws):
    for row in range(2, 11):
        ws.cell(row=row, column=1).number_format = FORMAT_DATE_XLSX14


# D: built-in via numFmtId lookup
def style_builtin_id(wb, ws):
    fmt_id = builtin_format_id(FORMAT_DATE_XLSX14)
    print("builtin_format_id for mm-dd-yy:", fmt_id)
    for row in range(2, 11):
        ws.cell(row=row, column=1).number_format = FORMAT_DATE_XLSX14


# E: validation only, no pre-formatting
def style_none(wb, ws):
    pass


# F: builtin + placeholder date value in first cell
def style_builtin_with_value(wb, ws):
    for row in range(2, 11):
        c = ws.cell(row=row, column=1)
        c.number_format = FORMAT_DATE_XLSX14
    ws.cell(row=2, column=1, value=date.today())


for label, fn in [
    ("named_style_yyyy_mm_dd", style_named),
    ("custom_yyyy_mm_dd", style_custom),
    ("builtin_mm_dd_yy", style_builtin),
    ("validation_only", style_none),
    ("builtin_with_today", style_builtin_with_value),
]:
    inspect(label, make_wb(fn))

for show_dd in [True, False]:
    inspect(f"builtin_showDropDown={show_dd}", make_wb(style_builtin, show_drop_down=show_dd))
