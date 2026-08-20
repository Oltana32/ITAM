"""Report generation services for CSV exports."""

import csv
import io

from django.utils import timezone

from apps.assets.models import Asset, AssetStatusHistory
from apps.assignments.models import Assignment
from apps.maintenance.models import MaintenanceRecord
from apps.licenses.models import SoftwareLicense


def _write_csv(headers: list[str], rows: list[list]) -> io.BytesIO:
    """Write CSV data to a BytesIO buffer."""
    text_buffer = io.StringIO()
    writer = csv.writer(text_buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output = io.BytesIO(text_buffer.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return output


def generate_asset_report() -> io.BytesIO:
    """Generate Asset Report: Asset Tag, Asset Name, Category, Status, Location."""
    headers = ["Asset Tag", "Asset Name", "Category", "Status", "Location", "Purchase Cost", "Depreciated Amount", "Current Value"]
    assets = Asset.objects.select_related("location").all().order_by("tag")
    rows = []
    for asset in assets:
        depr = asset.calculate_depreciation() or {}
        rows.append([
            asset.tag,
            asset.name,
            asset.get_category_display(),
            asset.get_status_display(),
            asset.location.name,
            getattr(asset.purchase_cost, 'quantize', lambda x: asset.purchase_cost) if asset.purchase_cost is not None else '',
            depr.get('depreciated_value', ''),
            depr.get('current_value', ''),
        ])
    return _write_csv(headers, rows)


def generate_assignment_report() -> io.BytesIO:
    """Generate Assignment Report: Asset, Assignee, Assignment Date, Return Date."""
    headers = [
        "Asset Tag",
        "Asset Name",
        "Assigned To",
        "Employee ID",
        "Assignment Date",
        "Return Date",
        "Status",
    ]
    assignments = Assignment.objects.select_related("asset").all().order_by("-assigned_date")
    rows = [
        [
            assignment.asset.tag,
            assignment.asset.name,
            assignment.assigned_to_name,
            assignment.employee_id,
            assignment.assigned_date,
            assignment.actual_return_date,
            assignment.get_status_display(),
        ]
        for assignment in assignments
    ]
    return _write_csv(headers, rows)


def generate_maintenance_report() -> io.BytesIO:
    """Generate Maintenance Report: Asset, Work Order ID, Maintenance Status, Cost."""
    headers = [
        "Asset Tag",
        "Asset Name",
        "Work Order ID",
        "Type",
        "Status",
        "Scheduled Date",
        "Cost",
    ]
    maintenance = MaintenanceRecord.objects.select_related("asset").all().order_by("-schedule_date")
    rows = [
        [
            record.asset.tag,
            record.asset.name,
            f"MNT-{record.id:06d}",
            record.get_type_display(),
            record.get_status_display(),
            record.schedule_date,
            record.cost,
        ]
        for record in maintenance
    ]
    return _write_csv(headers, rows)


def generate_license_report() -> io.BytesIO:
    """Generate License Report: Software, Vendor, Expiry Date, Seats."""
    headers = [
        "Software Name",
        "Vendor",
        "Seats",
        "Allocated Seats",
        "Available Seats",
        "Expiry Date",
        "Status",
        "Annual Cost",
    ]
    licenses = SoftwareLicense.objects.all().order_by("software_name")
    today = timezone.now().date()

    rows = []
    for license in licenses:
        if license.expiry_date:
            if license.expiry_date < today:
                status = "Expired"
            elif (license.expiry_date - today).days <= 30:
                status = "Expiring Soon"
            else:
                status = "Active"
        else:
            status = "No Expiry Date"

        rows.append([
            license.software_name,
            license.get_vendor_display(),
            license.seats,
            license.allocated_seats,
            license.available_seats,
            license.expiry_date,
            status,
            license.annual_cost,
        ])

    return _write_csv(headers, rows)


def generate_status_history_report() -> io.BytesIO:
    """Generate Asset Change History Report."""
    from apps.users.models import UserRole

    headers = [
        "Asset Tag",
        "Asset Name",
        "Change Type",
        "Field",
        "Old Value",
        "New Value",
        "Changed By",
        "Role",
        "Changed At",
        "Reason",
    ]
    history = (
        AssetStatusHistory.objects
        .select_related("asset", "changed_by")
        .order_by("-changed_at")
    )

    rows = []
    for entry in history:
        user = entry.changed_by
        if user:
            name = f"{user.first_name} {user.last_name}".strip() or user.email
            role = dict(UserRole.choices).get(getattr(user, "role", ""), "")
        else:
            name = "System"
            role = ""

        field_label = entry.field_name.replace("_", " ").title() if entry.field_name else ""

        rows.append([
            entry.asset.tag,
            entry.asset.name,
            entry.get_change_type_display(),
            field_label,
            entry.old_value,
            entry.new_value,
            name,
            role,
            entry.changed_at.strftime("%Y-%m-%d %H:%M:%S"),
            entry.reason,
        ])

    return _write_csv(headers, rows)


def generate_audit_report(audit_id: int) -> io.BytesIO:
    """Generate an Excel (.xlsx) report for a specific audit session.

    Sheet1: Summary
    Sheet2: Findings (one row per audited asset)
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from apps.audits.models import AuditSession

    session = AuditSession.objects.filter(pk=audit_id).prefetch_related('findings__asset', 'findings__auditor').first()
    if not session:
        raise ValueError(f"Audit session with id={audit_id} not found")

    wb = Workbook()
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Audit ID", session.audit_id),
        ("Title", session.title),
        ("Type", session.get_audit_type_display()),
        ("Status", session.get_status_display()),
        ("Lead Auditor", str(session.lead_auditor) if session.lead_auditor else ""),
        ("Planned Date", str(session.planned_date)),
        ("Audit Date", str(session.audit_date or "")),
        ("Location", str(session.location) if session.location else ""),
        ("Department", str(session.department) if session.department else ""),
        ("Total Expected", session.total_assets_audited),
        ("Total Found", session.assets_found),
        ("Total Missing", session.assets_not_found),
        ("Assets With Issues", session.assets_with_issues),
        ("Created At", session.created_at.strftime("%Y-%m-%d %H:%M:%S") if session.created_at else ""),
    ]

    for r_idx, (k, v) in enumerate(summary_rows, start=1):
        ws.cell(row=r_idx, column=1, value=k)
        ws.cell(row=r_idx, column=2, value=v)

    # Findings sheet
    ws2 = wb.create_sheet(title="Findings")
    headers = [
        "Asset Tag",
        "Asset Name",
        "Expected Status",
        "Finding Status",
        "Result",
        "Auditor",
        "Verified At",
        "Current Location",
        "Current Condition",
        "Notes",
        "Verification JSON",
    ]
    for c_idx, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=c_idx, value=h)

    for r_idx, finding in enumerate(session.findings.all().order_by('-verified_at'), start=2):
        asset = getattr(finding, 'asset', None)
        ws2.cell(row=r_idx, column=1, value=getattr(asset, 'tag', ''))
        ws2.cell(row=r_idx, column=2, value=getattr(asset, 'name', ''))
        # expected status is not tracked on session; leave blank or map from asset
        ws2.cell(row=r_idx, column=3, value='')
        ws2.cell(row=r_idx, column=4, value=finding.get_status_display())
        ws2.cell(row=r_idx, column=5, value=finding.result_status)
        ws2.cell(row=r_idx, column=6, value=str(finding.auditor) if finding.auditor else '')
        ws2.cell(row=r_idx, column=7, value=finding.verified_at.strftime("%Y-%m-%d %H:%M:%S") if finding.verified_at else '')
        ws2.cell(row=r_idx, column=8, value=str(finding.current_location) if finding.current_location else '')
        ws2.cell(row=r_idx, column=9, value=finding.current_condition)
        ws2.cell(row=r_idx, column=10, value=finding.notes)
        # verification as compact JSON string
        try:
            import json

            verification_str = json.dumps(finding.verification or {})
        except Exception:
            verification_str = str(finding.verification or "")
        ws2.cell(row=r_idx, column=11, value=verification_str)

    # Auto-adjust column widths for findings sheet
    for i, column_cells in enumerate(ws2.columns, start=1):
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        ws2.column_dimensions[get_column_letter(i)].width = min(max(length + 2, 10), 60)

    # Save workbook to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_audit_csv_report(audit_id: int) -> io.BytesIO:
    """Generate a CSV report for a specific audit session.

    Columns: Asset Tag, Asset Name, Finding Status, Result, Auditor, Verified At,
    Current Location, Current Condition, Notes, Verification JSON
    """
    from apps.audits.models import AuditSession
    import json

    session = AuditSession.objects.filter(pk=audit_id).prefetch_related('findings__asset', 'findings__auditor').first()
    if not session:
        raise ValueError(f"Audit session with id={audit_id} not found")

    headers = [
        "Asset Tag",
        "Asset Name",
        "Finding Status",
        "Result",
        "Auditor",
        "Verified At",
        "Current Location",
        "Current Condition",
        "Notes",
        "Verification JSON",
    ]

    rows: list[list] = []
    for finding in session.findings.all().order_by('-verified_at'):
        asset = getattr(finding, 'asset', None)
        rows.append([
            getattr(asset, 'tag', ''),
            getattr(asset, 'name', ''),
            finding.get_status_display(),
            finding.result_status,
            str(finding.auditor) if finding.auditor else '',
            finding.verified_at.strftime("%Y-%m-%d %H:%M:%S") if finding.verified_at else '',
            str(finding.current_location) if finding.current_location else '',
            finding.current_condition,
            finding.notes,
            json.dumps(finding.verification or {}),
        ])

    return _write_csv(headers, rows)
