"""Build the bulk-import Excel template with dropdowns and date validation."""

from __future__ import annotations

from datetime import date

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from apps.assets.models import AssetCategory, AssetCondition
from apps.assets.specs import CATEGORY_SPECS

BULK_IMPORT_BASE_HEADERS = [
    "Name",
    "Category",
    "Status",
    "Manufacturer",
    "Serial Number",
    "Model",
    "Purchase Date",
    "Purchase Cost",
    "Warranty Expiry",
    "Condition",
    "Department",
    "Notes",
    "Location",
]

BULK_IMPORT_SKIP_SHEETS = {"categories", "specs", "instructions", "_lists"}
BULK_IMPORT_ALL_CATEGORIES_SHEET = "all categories"

# Match the add-asset form: hide ready/assigned from dropdown choices.
IMPORT_TEMPLATE_STATUSES = [
    "available",
    "in-use",
    "maintenance",
    "retired",
    "disposed",
    "lost",
    "damaged",
]

MAX_TEMPLATE_DATA_ROWS = 500
LIST_SHEET_NAME = "_Lists"
# Excel desktop expects native date formulas for date validation so the picker
# can appear when a user selects the date cells.
DATE_VALIDATION_MIN = "DATE(2000,1,1)"
DATE_VALIDATION_MAX = "DATE(2100,12,31)"
DATE_STYLE_NAME = "itam_import_date"
DATE_COLUMN_WIDTH = 14


class _ListRegistry:
    """Store dropdown option lists on a hidden worksheet."""

    def __init__(self, lists_ws):
        self.lists_ws = lists_ws
        self._next_col = 1
        self._ranges: dict[str, str] = {}

    def register(self, key: str, values: list[str]) -> str:
        if key in self._ranges:
            return self._ranges[key]

        col = self._next_col
        for row_idx, value in enumerate(values, start=1):
            self.lists_ws.cell(row=row_idx, column=col, value=value)

        col_letter = get_column_letter(col)
        range_ref = f"{quote_sheetname(LIST_SHEET_NAME)}!${col_letter}$1:${col_letter}${len(values)}"
        self._ranges[key] = range_ref
        self._next_col += 1
        return range_ref


def _header_index(headers: list, name: str) -> int | None:
    target = name.strip().lower()
    for idx, header in enumerate(headers, start=1):
        if header is not None and str(header).strip().lower() == target:
            return idx
    return None


def _parse_spec_key(header: str) -> str | None:
    if not header or not isinstance(header, str) or "(spec:" not in header:
        return None
    start = header.find("(spec:") + len("(spec:")
    end = header.find(")", start)
    key = header[start:end].strip() if end != -1 else header[start:].strip()
    return key or None


def _ensure_date_style(wb: openpyxl.Workbook) -> None:
    """Register a workbook date style so Excel treats cells as dates (shows calendar picker)."""
    if DATE_STYLE_NAME in wb.named_styles:
        return

    date_style = NamedStyle(name=DATE_STYLE_NAME)
    date_style.number_format = FORMAT_DATE_YYYYMMDD2
    wb.add_named_style(date_style)


def _apply_list_validation(ws, col_idx: int, range_ref: str, prompt: str) -> None:
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{MAX_TEMPLATE_DATA_ROWS}"
    dv = DataValidation(
        type="list",
        formula1=f"={range_ref}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        showInputMessage=True,
    )
    dv.error = "Please choose a value from the dropdown list."
    dv.errorTitle = "Invalid value"
    dv.prompt = prompt
    dv.promptTitle = "Select a value"
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _apply_date_validation(ws, col_idx: int) -> None:
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{MAX_TEMPLATE_DATA_ROWS}"
    # Excel date-validation formulas must use serial numbers. ISO strings like
    # "2000-01-01" are parsed as arithmetic (2000 - 1 - 1) and break validation.
    # Combined with yyyy-mm-dd cell formatting, Excel 365 / Excel Online show a
    # native calendar icon when the cell is selected.
    dv = DataValidation(
        type="date",
        operator="between",
        formula1=DATE_VALIDATION_MIN,
        formula2=DATE_VALIDATION_MAX,
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorStyle="stop",
    )
    dv.error = "Enter a valid date (YYYY-MM-DD)."
    dv.errorTitle = "Invalid date"
    dv.prompt = "Click the calendar icon or type a date (YYYY-MM-DD)."
    dv.promptTitle = "Select a date"
    dv.add(cell_range)
    ws.add_data_validation(dv)

    ws.column_dimensions[col_letter].width = DATE_COLUMN_WIDTH
    for row in range(2, MAX_TEMPLATE_DATA_ROWS + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.style = DATE_STYLE_NAME
        cell.number_format = "yyyy-mm-dd"


def _apply_sheet_validations(ws, headers: list, list_registry: _ListRegistry, sheet_category: str | None) -> None:
    category_range = list_registry.register(
        "categories",
        [category.value for category in AssetCategory],
    )
    status_range = list_registry.register("statuses", IMPORT_TEMPLATE_STATUSES)
    condition_range = list_registry.register(
        "conditions",
        [condition.value for condition in AssetCondition],
    )

    category_col = _header_index(headers, "Category")
    if category_col:
        _apply_list_validation(ws, category_col, category_range, "Choose an asset category.")

    status_col = _header_index(headers, "Status")
    if status_col:
        _apply_list_validation(ws, status_col, status_range, "Choose an asset status.")

    condition_col = _header_index(headers, "Condition")
    if condition_col:
        _apply_list_validation(ws, condition_col, condition_range, "Choose asset condition.")

    purchase_date_col = _header_index(headers, "Purchase Date")
    if purchase_date_col:
        _apply_date_validation(ws, purchase_date_col)

    warranty_col = _header_index(headers, "Warranty Expiry")
    if warranty_col:
        _apply_date_validation(ws, warranty_col)

    category_specs = CATEGORY_SPECS.get(sheet_category or "", {})
    spec_fields = category_specs.get("fields", {})
    for col_idx, header in enumerate(headers, start=1):
        spec_key = _parse_spec_key(header if isinstance(header, str) else "")
        if not spec_key:
            continue
        field_cfg = spec_fields.get(spec_key, {})
        if field_cfg.get("type") == "date":
            _apply_date_validation(ws, col_idx)
            continue
        options = field_cfg.get("options")
        if field_cfg.get("type") != "select" or not options:
            continue
        list_key = f"spec:{spec_key}"
        options_range = list_registry.register(list_key, list(options))
        label = field_cfg.get("label") or spec_key
        _apply_list_validation(ws, col_idx, options_range, f"Choose {label.lower()}.")


def build_import_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    _ensure_date_style(wb)
    lists_ws = wb.create_sheet(LIST_SHEET_NAME)
    lists_ws.sheet_state = "hidden"
    list_registry = _ListRegistry(lists_ws)

    base_headers = [header for header in BULK_IMPORT_BASE_HEADERS if header != "Category"]
    all_categories_headers = list(BULK_IMPORT_BASE_HEADERS)

    ws = wb.active
    ws.title = "All Categories"
    ws.append(all_categories_headers)
    _apply_sheet_validations(ws, all_categories_headers, list_registry, sheet_category=None)

    instructions = wb.create_sheet(title="Instructions")
    instructions.append(["Bulk import guide"])
    instructions.append([""])
    instructions.append(["Use one sheet per asset category (laptop, desktop, monitor, etc.)"])
    instructions.append(["Each category sheet includes the same fields as the add-asset form for that category"])
    instructions.append(["Required columns: Name, Manufacturer, Serial Number, Model, Purchase Date"])
    instructions.append(["Category, Status, Condition, and spec fields use dropdown lists like the asset form"])
    instructions.append(["Purchase Date and Warranty Expiry are Excel date cells — select a cell and Excel desktop will show the calendar picker"])
    instructions.append(["Location is optional — if blank, the first location in the system is used"])
    instructions.append(["Manufacturer is auto-created if it does not already exist"])
    instructions.append(["Delete any sample rows before importing"])

    ws2 = wb.create_sheet(title="Categories")
    ws2.append(["Valid categories:"])
    for category in AssetCategory:
        ws2.append([category.value])

    for cat_key, cat_cfg in CATEGORY_SPECS.items():
        if cat_key not in {category.value for category in AssetCategory}:
            continue
        sheet = wb.create_sheet(title=cat_key)
        headers = list(base_headers)
        fields = cat_cfg.get("fields", {})
        for field_key, field_cfg in fields.items():
            label = field_cfg.get("label") or field_key
            headers.append(f"{label} (spec:{field_key})")
        sheet.append(headers)
        _apply_sheet_validations(sheet, headers, list_registry, sheet_category=cat_key)

    ws3 = wb.create_sheet(title="Specs")
    ws3.append(["Category", "Field Key", "Field Label", "Type", "Options (comma-separated)"])
    for cat_key, cat_cfg in CATEGORY_SPECS.items():
        fields = cat_cfg.get("fields", {})
        if not fields:
            ws3.append([cat_key, "", "", "", ""])
            continue
        for field_key, field_cfg in fields.items():
            options = field_cfg.get("options")
            options_str = ", ".join(options) if options else ""
            ws3.append([cat_key, field_key, field_cfg.get("label", ""), field_cfg.get("type", ""), options_str])

    return wb
