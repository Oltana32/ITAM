from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient

import openpyxl
import io
from datetime import date

from apps.assets.models import Asset
from apps.manufacturers.models import Manufacturer
from apps.locations.models import Location
from apps.users.models import UserRole


User = get_user_model()

ALL_CATEGORIES_HEADERS = [
    "Name",
    "Category",
    "Status",
    "Manufacturer",
    "Serial Number",
    "Model",
    "Purchase Date",
    "Purchase Cost",
    "Warranty Expiry",
    "Condition",
    "Department",
    "Notes",
    "Location",
]


class AssetBulkImportTests(TestCase):
    def setUp(self):
        self.manufacturer, _ = Manufacturer.objects.get_or_create(name="MfgCo")
        self.location = Location.objects.create(name="Main Office")
        self.user = User.objects.create_user(email="admin@bulk.com", password="pass123", role=UserRole.ADMIN)
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def make_all_categories_workbook(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Categories"
        ws.append(ALL_CATEGORIES_HEADERS)
        for row in rows:
            ws.append(row)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def make_category_workbook(self, category, rows, spec_headers=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = category
        headers = [h for h in ALL_CATEGORIES_HEADERS if h != "Category"]
        if spec_headers:
            headers.extend(spec_headers)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def post_workbook(self, bio):
        wb_file = ("import.xlsx", bio, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return self.client.post("/api/assets/bulk-import/", {"file": wb_file}, format="multipart")

    def test_bulk_import_happy_path_and_errors(self):
        rows = [
            ["Laptop A", "laptop", "available", "MfgCo", "SN-001", "X", "2026-01-01", 1000, "2028-01-01", "good", "IT", "notes", "Main Office"],
            ["Phone B", "phone", "available", "MfgCo", "", "P1", "2026-02-01", 500, "2028-02-01", "good", "Sales", "notes", ""],
            ["BadCat", "not-a-cat", "available", "MfgCo", "SN-003", "M", "2026-03-01", None, None, "good", "Ops", "notes", ""],
        ]
        response = self.post_workbook(self.make_all_categories_workbook(rows))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["summary"]["created"], 1)
        self.assertEqual(data["summary"]["failed"], 2)

    def test_auto_creates_manufacturer_and_uses_default_location(self):
        rows = [
            ["New Asset", "laptop", "available", "Brand New Mfg", "SN-NEW", "Model Z", "2026-01-01", 100, None, "good", "", "", ""],
        ]
        response = self.post_workbook(self.make_all_categories_workbook(rows))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["summary"]["created"], 1)
        asset = Asset.objects.get(serial_number="SN-NEW")
        self.assertEqual(asset.manufacturer.name, "Brand New Mfg")
        self.assertEqual(asset.location_id, self.location.id)

    def test_category_sheet_import_with_specs(self):
        rows = [
            [
                "Office Laptop",
                "available",
                "MfgCo",
                "SN-LAP-1",
                "Latitude",
                "2026-01-01",
                1200,
                None,
                "good",
                "IT",
                "",
                "",
                "Intel Core i5",
                "16GB",
                "512GB SSD",
                "Windows 11 Pro",
            ],
        ]
        bio = self.make_category_workbook(
            "laptop",
            rows,
            spec_headers=[
                "CPU (spec:cpu)",
                "RAM (spec:ram)",
                "Storage (spec:storage)",
                "Operating System (spec:os)",
            ],
        )
        response = self.post_workbook(bio)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["summary"]["created"], 1)
        asset = Asset.objects.get(serial_number="SN-LAP-1")
        self.assertEqual(asset.category, "laptop")
        self.assertEqual(asset.specs["cpu"], "Intel Core i5")
        self.assertEqual(asset.specs["ram"], "16GB")

    def test_skips_reference_sheets(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Categories"
        ws.append(["Valid categories:"])
        ws.append(["laptop"])
        ws2 = wb.create_sheet("Specs")
        ws2.append(["Category", "Field Key"])
        ws2.append(["laptop", "cpu"])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        response = self.post_workbook(bio)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["summary"]["created"], 0)

    def test_import_template_includes_dropdown_and_date_validation(self):
        response = self.client.get("/api/assets/import-template/")
        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        self.assertIn("_Lists", wb.sheetnames)
        self.assertEqual(wb["_Lists"].sheet_state, "hidden")

        all_categories = wb["All Categories"]
        validations = list(all_categories.data_validations.dataValidation)
        validation_types = {dv.type for dv in validations}
        self.assertIn("list", validation_types)
        self.assertIn("date", validation_types)

        date_validations = [dv for dv in validations if dv.type == "date"]
        self.assertGreaterEqual(len(date_validations), 2)
        expected_min = "DATE(2000,1,1)"
        expected_max = "DATE(2100,12,31)"
        for dv in date_validations:
            self.assertEqual(str(dv.formula1), expected_min)
            self.assertEqual(str(dv.formula2), expected_max)
            self.assertEqual(dv.promptTitle, "Select a date")

        purchase_date_cell = all_categories.cell(row=2, column=ALL_CATEGORIES_HEADERS.index("Purchase Date") + 1)
        self.assertEqual(purchase_date_cell.number_format, "yyyy-mm-dd")

        laptop_sheet = wb["laptop"]
        laptop_validations = list(laptop_sheet.data_validations.dataValidation)
        laptop_types = {dv.type for dv in laptop_validations}
        self.assertIn("list", laptop_types)
        self.assertIn("date", laptop_types)
        self.assertGreaterEqual(len(laptop_validations), 6)
