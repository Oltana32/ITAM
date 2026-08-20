from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response

from django_filters import rest_framework as filters

from apps.locations.models import Location
from config.permissions import IsITStaffOrAdmin, IsITStaffOrAdminOrReadOnly

from .history import log_asset_creation, log_asset_updates, log_status_change
from .models import Asset, AssetStatusHistory
from .serializers import AssetSerializer, AssetStatusHistorySerializer
from .services import retire_asset, set_asset_location
from apps.notifications.services import notify_asset_retired, notify_asset_disposed
from .specs import CATEGORY_SPECS
from .models import AssetCategory
from .import_template import (
    BULK_IMPORT_ALL_CATEGORIES_SHEET,
    BULK_IMPORT_SKIP_SHEETS,
    build_import_workbook,
)
from apps.core.constants import AssetStatus
from apps.manufacturers.models import Manufacturer
import openpyxl
import io
from django.db import transaction
from django.http import HttpResponse


class AssetFilter(filters.FilterSet):
    status = filters.CharFilter(field_name="status", lookup_expr="exact")
    category = filters.CharFilter(field_name="category", lookup_expr="exact")
    manufacturer = filters.NumberFilter(field_name="manufacturer_id", lookup_expr="exact")
    location = filters.NumberFilter(field_name="location_id", lookup_expr="exact")

    class Meta:
        model = Asset
        fields = ["status", "category", "manufacturer", "location"]


class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.select_related(
        "manufacturer", "location", "created_by", "updated_by"
    ).all()
    serializer_class = AssetSerializer
    permission_classes = [IsITStaffOrAdminOrReadOnly]
    filter_backends = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class = AssetFilter
    search_fields = ("name", "tag", "serial_number", "model", "department")
    ordering_fields = ("name", "tag", "status", "created_at", "updated_at")

    def get_queryset(self):
        qs = super().get_queryset()
        include_retired = self.request.query_params.get("include_retired", "false").lower() in (
            "1",
            "true",
            "yes",
        )
        if not include_retired:
            qs = qs.exclude(status__in=("retired", "disposed"))
        return qs

    def perform_create(self, serializer):
        asset = serializer.save(created_by=self.request.user)
        log_asset_creation(asset, changed_by=self.request.user)

    def perform_update(self, serializer):
        old_asset = Asset.objects.select_related("manufacturer", "location").get(
            pk=serializer.instance.pk
        )
        asset = serializer.save(updated_by=self.request.user)
        if old_asset.status != asset.status:
            if asset.status == AssetStatus.RETIRED:
                notify_asset_retired(asset)
            elif asset.status == AssetStatus.DISPOSED:
                notify_asset_disposed(asset)
        log_asset_updates(asset, old_asset, changed_by=self.request.user)

    @action(detail=True, methods=["get"])
    def history(self, request, pk=None):
        """GET /api/assets/:id/history/ — change log for this asset."""
        asset = self.get_object()
        qs = (
            AssetStatusHistory.objects.filter(asset=asset)
            .select_related("changed_by")
            .order_by("-changed_at")
        )
        serializer = AssetStatusHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], permission_classes=[IsITStaffOrAdmin])
    def move(self, request, pk=None):
        """POST /api/assets/:id/move/ { "location": <id> }"""
        asset = self.get_object()
        loc_id = request.data.get("location")
        if not loc_id:
            return Response({"location": "Required."}, status=status.HTTP_400_BAD_REQUEST)
        location = Location.objects.filter(pk=loc_id).first()
        if not location:
            return Response({"location": "Invalid location."}, status=status.HTTP_400_BAD_REQUEST)

        old_asset = Asset.objects.select_related("manufacturer", "location").get(pk=asset.pk)
        set_asset_location(asset, location)
        asset.refresh_from_db()
        log_asset_updates(asset, old_asset, changed_by=request.user, fields=("location",))
        return Response(AssetSerializer(asset).data)

    @action(detail=True, methods=["post"], permission_classes=[IsITStaffOrAdmin])
    def retire(self, request, pk=None):
        asset = self.get_object()
        old_status = asset.status
        retire_asset(asset)
        asset.refresh_from_db()
        log_status_change(
            asset,
            old_status,
            asset.status,
            changed_by=request.user,
            reason="Asset retired",
        )
        notify_asset_retired(asset)
        return Response(AssetSerializer(asset).data)

    @action(detail=False, methods=["get"])
    def specs_config(self, request):
        """GET /api/assets/specs_config/ — get all category specs configurations."""
        return Response(CATEGORY_SPECS)

    @action(detail=False, methods=["get"])
    def category_specs(self, request):
        """GET /api/assets/category_specs/?category=laptop — get specs for specific category."""
        category = request.query_params.get("category")
        if not category:
            return Response({"error": "category parameter is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        specs = CATEGORY_SPECS.get(category, {})
        if not specs:
            return Response({"error": f"No specs configuration found for category: {category}"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response(specs)

    def _resolve_bulk_import_location(self, location_name):
        """Match add-asset form behavior: optional location, fall back to first location."""
        if location_name and str(location_name).strip():
            location = Location.objects.filter(name__iexact=str(location_name).strip()).first()
            if location:
                return location, None
        default_location = Location.objects.order_by("id").first()
        if default_location:
            return default_location, None
        return None, "No locations exist. Ask an admin to add a location first."

    def _resolve_bulk_import_manufacturer(self, manufacturer_name):
        """Match add-asset form behavior: auto-create manufacturer when missing."""
        mfg_name = str(manufacturer_name or "").strip()
        if not mfg_name:
            return None, "Manufacturer is required"
        manufacturer = Manufacturer.objects.filter(name__iexact=mfg_name).first()
        if not manufacturer:
            manufacturer = Manufacturer.objects.create(name=mfg_name)
        return manufacturer, None

    @action(detail=False, methods=["post"], permission_classes=[IsITStaffOrAdmin], url_path="bulk-import")
    def bulk_import(self, request):
        """POST /api/assets/bulk-import/ — upload an .xlsx file and import rows as assets.

        Returns per-row created/errors and a summary.
        """
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "file is required"}, status=400)

        # Basic limits
        if uploaded.size > 5 * 1024 * 1024:
            return Response({"detail": "File too large (max 5MB)"}, status=400)

        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(uploaded.read()), data_only=True)
        except Exception as e:
            return Response({"detail": f"Failed to read Excel file: {e}"}, status=400)

        created = []
        errors = []
        seen_serials = set()
        total_rows = 0
        max_rows = 2000
        valid_categories = {c.value for c in AssetCategory}
        valid_statuses = {c.value for c in AssetStatus}

        def _normalize(h):
            return str(h).strip().lower() if h is not None else ""

        for sheet in wb.worksheets:
            sheet_title = (sheet.title or "").strip().lower()
            if sheet_title in BULK_IMPORT_SKIP_SHEETS:
                continue
            if sheet_title != BULK_IMPORT_ALL_CATEGORIES_SHEET and sheet_title not in valid_categories:
                continue

            try:
                header_row = next(sheet.iter_rows(min_row=1, max_row=1))
            except StopIteration:
                continue

            headers = [cell.value for cell in header_row]
            header_map = {_normalize(h): idx for idx, h in enumerate(headers)}

            spec_cols = {}
            for idx, h in enumerate(headers):
                if not h or not isinstance(h, str) or "(spec:" not in h:
                    continue
                start = h.find("(spec:") + len("(spec:")
                end = h.find(")", start)
                key = h[start:end].strip() if end != -1 else h[start:].strip()
                if key:
                    spec_cols[key] = idx

            sheet_category = sheet_title if sheet_title in valid_categories else None

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                def get_col(name):
                    idx = header_map.get(name)
                    if idx is None:
                        return None
                    return row[idx].value

                name = get_col("name")
                serial = get_col("serial number")
                category = sheet_category or (get_col("category") or "")
                if isinstance(category, str):
                    category = category.strip().lower()
                if category == "printer":
                    category = "equipment"

                if not (name or serial):
                    continue

                if total_rows >= max_rows:
                    return Response({"detail": f"Too many rows (limit {max_rows})"}, status=400)
                total_rows += 1

                manufacturer_name = get_col("manufacturer") or ""
                location_name = get_col("location") or ""
                condition = get_col("condition") or None
                model = get_col("model") or ""
                purchase_date = get_col("purchase date")
                purchase_cost = get_col("purchase cost")
                warranty_expiry = get_col("warranty expiry")
                department = get_col("department") or ""
                notes = get_col("notes") or ""
                status_raw = get_col("status") or "available"
                status = str(status_raw).strip().lower().replace("-", "_") if status_raw else "available"

                row_errors = {}
                if not name or not str(name).strip():
                    row_errors["name"] = "Name is required"

                if not category or category not in valid_categories:
                    row_errors["category"] = f'Invalid category. Valid: {", ".join(sorted(valid_categories))}'

                if status not in valid_statuses:
                    row_errors["status"] = f'Invalid status. Valid: {", ".join(sorted(valid_statuses))}'

                manufacturer, manufacturer_error = self._resolve_bulk_import_manufacturer(manufacturer_name)
                if manufacturer_error:
                    row_errors["manufacturer"] = manufacturer_error

                location, location_error = self._resolve_bulk_import_location(location_name)
                if location_error:
                    row_errors["location"] = location_error

                serial_text = str(serial).strip() if serial is not None else ""
                if not serial_text:
                    row_errors["serial_number"] = "Serial number is required"
                elif serial_text.lower() in seen_serials:
                    row_errors["serial_number"] = "Duplicate serial number in file"
                elif Asset.objects.filter(serial_number__iexact=serial_text).exists():
                    row_errors["serial_number"] = "Serial number already exists"

                if not model or not str(model).strip():
                    row_errors["model"] = "Model is required"

                if purchase_date in (None, ""):
                    row_errors["purchase_date"] = "Purchase date is required"

                if row_errors:
                    errors.append({"row": f"{sheet.title}:{row_idx}", "errors": row_errors})
                    continue

                specs = {}
                for key, idx in spec_cols.items():
                    val = row[idx].value if idx < len(row) else None
                    if val is not None and val != "":
                        specs[key] = val

                for possible_key in header_map:
                    if possible_key.startswith("spec:"):
                        continue
                    if possible_key in CATEGORY_SPECS.get(category, {}).get("fields", {}):
                        idx = header_map.get(possible_key)
                        if idx is not None and idx < len(row):
                            val = row[idx].value
                            if val is not None and val != "":
                                specs[possible_key] = val

                try:
                    with transaction.atomic():
                        asset = Asset.objects.create(
                            name=str(name).strip(),
                            category=category,
                            status=status,
                            manufacturer=manufacturer,
                            location=location,
                            serial_number=serial_text,
                            condition=str(condition).strip().lower() if condition else "good",
                            model=str(model).strip(),
                            purchase_date=purchase_date,
                            purchase_cost=purchase_cost,
                            warranty_expiry=warranty_expiry,
                            department=str(department) if department else "",
                            notes=str(notes) if notes else "",
                            specs=specs,
                            created_by=request.user,
                        )
                        log_asset_creation(asset, changed_by=request.user)
                        created.append({"row": f"{sheet.title}:{row_idx}", "tag": asset.tag, "id": asset.id})
                        seen_serials.add(serial_text.lower())
                except Exception as e:
                    errors.append({"row": f"{sheet.title}:{row_idx}", "errors": {"exception": str(e)}})

        summary = {"total_rows": total_rows, "created": len(created), "failed": len(errors)}
        return Response({"created": created, "errors": errors, "summary": summary})

    @action(detail=False, methods=["get"], permission_classes=[IsITStaffOrAdminOrReadOnly], url_path="import-template")
    def import_template(self, request):
        """Return an .xlsx template for imports."""
        wb = build_import_workbook()
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="asset_import_template.xlsx"'
        return resp
